import xlsxwriter
import openpyxl
from os.path import exists
import pandas as pd
# Workbook() takes one, non-optional, argument
xlname='new.xlsx'
def create():#create xls if its not present
    if not exists(xlname):
    	workbook = xlsxwriter.Workbook(xlname)
    	wb=openpyxl.Workbook()
    	wbook=wb.active
    	wbook.title="object detection"
    	wbook.append(("ip adress",'  ',"objects detected"))
    	print("file created")
    	wb.save(xlname)
    	wb.close()
    else:
        print("file already exists")
def sip(i=""):#put value to xls element
    workbook = xlsxwriter.Workbook(xlname)
    wb=openpyxl.Workbook()
    wbook=wb.active
    wbook['A2']=i
    wb.save(xlname)
    wb.close()
def update():#merge xls to main database
    ipdf=pd.read_excel("data.xlsx")
    spdf=pd.read_excel("new.xlsx")
    ipdf.append(spdf, ignore_index=True)
    ipdf.to_excel("data.xlsx", index=False)
    print(ipdf)
    print("saved  before exit")
def grip():#get value of xls as string
    wb=openpyxl.load_workbook(xlname,data_only=True)
    sh=wb.active
    srt=sh['a2'].value
    wb.close()
    return srt
